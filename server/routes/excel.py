from flask import Blueprint, session, jsonify
from app import db

from sqlalchemy import text
import numpy as np
from flask_cors import cross_origin
import os
from openpyxl import load_workbook
from datetime import datetime, date
import random

excel_routes = Blueprint("excel_routes", __name__)


def get_piezometer_data():
    userQuery = db.session.execute(
        text(
            f"SELECT paddock as piezo_paddock ,id as piezo_name, depth as piezo_depth, datalogger as piezo_node, channel as piezo_channel FROM piezometer_details WHERE status = 1;"
        )
    )

    piezo_data = [dict(r._mapping) for r in userQuery]

    return piezo_data


def query(data_obj, ammount, period):
    userQuery = db.session.execute(
        text(
            f"select min(pressure) as min ,max(pressure) as max ,avg(pressure) as avg from node_{data_obj['piezo_node']}_{data_obj['piezo_channel']} where (current_date - time) <= interval '{ammount}' {period};"
        )
    )
    piezo_data = [dict(r._mapping) for r in userQuery]

    data_dict = piezo_data[0]

    return [data_dict["min"], data_dict["max"], data_dict["avg"]]


def get_data():
    piezo_data = get_piezometer_data()
    new_data = []
    # looping around all piezometers
    for data in piezo_data:
        weekly = query(data, 14, "day")
        monthly = query(data, 1, "month")
        quarterly = query(data, 3, "month")

        lectures_data = [weekly, monthly, quarterly]

        flat_list = list(np.concatenate(lectures_data).flat)

        new_data.append(
            {
                "piezo_depth": data["piezo_depth"],
                "piezo_name": data["piezo_name"],
                "piezo_paddock": data["piezo_paddock"],
                "piezo_lectures": flat_list,
            }
        )

    return new_data


@excel_routes.route("/api/v1/excel-data", methods=["GET"])
@cross_origin()
def get_excel_data():
    # tdata = get_data()
    print(os.path.abspath("pyreport/report.xlsx"))
    data = get_data()

    return jsonify(
        {
            # "excel-data": data,
            "today": date.today(),
            "path_to_empty_file": os.path.abspath("pyreport/report.xlsx"),
            "path_to_target": os.path.abspath("../client/public/pyreport/report2.xlsx"),
        }
    )


def read_excel():
    tdata = get_data()

    filename = os.path.abspath("pyreport/report.xlsx")
    # filename = BASEPATH + "/pyreport/report.xlsx"
    print(filename)

    wb = load_workbook(filename)

    sh = wb.active
    i = 14
    j = 7
    for data in tdata:
        sh.cell(row=i, column=3).value = data["piezo_paddock"]
        sh.cell(row=i, column=5).value = data["piezo_name"]
        sh.cell(row=i, column=16).value = data["piezo_depth"]

        for val in data["piezo_lectures"]:
            sh.cell(row=i, column=j).value = "-" if val is None else float(val)
            j += 1

        # for j in range(3,12):
        #     print(i,j)
        #     sh.cell(row=i,column=4+j).value = float(data[j])
        j = 7
        i += 1

    print("today:", date.today())
    sh.cell(row=5, column=13).value = date.today()
    sh.cell(row=75, column=2).value = "test"
    sh.cell(row=75, column=3).value = random.randint(0, 100)

    wb.save(os.path.abspath("../client/public/pyreport/report2.xlsx"))
    return os.path.abspath("../client/public/pyreport/report2.xlsx")


@excel_routes.route("/api/v1/modify_excel", methods=["POST"])
@cross_origin()
def modify_excel():
    now = datetime.now()
    dt_string = now.strftime("%Y/%m/%d %H:%M:%S")
    data = read_excel()
    print("file saved on:", data)
    print("report download by %s at %s" % (session.get("user_id"), dt_string))
    return jsonify(
        {"filename": os.path.abspath("../client/public/pyreport/report2.xlsx")}
    )
    # return jsonify({"filename": os.path.abspath("pyreport/report2.xlsx")})
