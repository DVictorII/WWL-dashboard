import psycopg2
import os
from decimal import Decimal
os.environ['OPENBLAS_NUM_THREADS'] = '1'
from openpyxl import load_workbook
from datetime import date

#BASEPATH = "clients.wwlengineering.com/public_html/"
BASEPATH = os.getcwd()+"/"

dbname = 'wwlengineering_rossing'
user = 'wwlengineering_admin_rossing'
password = 'yf=R(qH7Q#fG'
host = 'localhost'
dev = True

def dbconnect():
    conn = psycopg2.connect(host=host,
                            database=dbname,
                            user=user,
                            password=password)
    return conn

def get_piezometer_data():
    con = dbconnect()
    cur = con.cursor()
    query = "select paddock,id as name, depth, datalogger as node, " +\
         "channel from piezometer_details where status = 1"
    cur.execute(query)
    con.commit()
    data = cur.fetchall()
    return data

def query(data,node,channel,ammount,period):
    query = "select min(pressure),max(pressure),avg(pressure) "+\
            "from node_%s_%s where (current_date - time) <= interval '%s' %s"%(node,channel,ammount,period)
    con = dbconnect()
    cur = con.cursor()
    cur.execute(query)
    con.commit()
    vals = cur.fetchall()
    for x in vals:
        data = data + tuple(map(str,list(x)))
    return data

def get_data():
    piezo_data = get_piezometer_data()
    new_data = []
    # looping around all piezometers
    for data in piezo_data:
        node,channel = data[3],data[4]
        data = data[:3]
        data = query(data,node,channel,14,'day')
        data = query(data,node,channel,1,'month')
        data = query(data,node,channel,3,'month')
        new_data.append(data)
    return new_data

def read_excel():
    print("#######################")
    tdata = get_data()
    filename = BASEPATH + "pyreport/report.xlsx"
    #print(piezo_data)
    wb = load_workbook(filename)
    sh = wb.active
    i=14
    for data in tdata:
        print(data)
        sh.cell(row=i,column=3).value = data[0]
        sh.cell(row=i,column=5).value = data[1]
        sh.cell(row=i,column=16).value = data[2]
        for j in range(3,12):
            print(i,j)
            sh.cell(row=i,column=4+j).value = float(data[j])
        i+=1
    sh.cell(row=5,column=13).value = date.today()
    wb.save(BASEPATH + "pyreport/report2.xlsx")
    return i

if __name__ == "__main__":
    print(read_excel())